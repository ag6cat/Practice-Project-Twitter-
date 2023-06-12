import os
import openpyxl


# add search word to record the result of a search
def creat_xlsx(results):
    csvfile = './Twitter_Data/Twitter.xlsx'
    #need name generator to avoid duplicate file, for now writing over a same file
    
    if os.path.exists(csvfile):
        workbook = openpyxl.load_workbook(csvfile)
    else:
        workbook = openpyxl.Workbook()
    save_file = csvfile
    worksheet = workbook.active
    max_row = worksheet.max_row
    
    for r in range(len(results)):
        for c in range(len(results[0])):
            try:
                worksheet.cell(r + max_row, c + 1).value = results[r][c]
            except:
                worksheet.cell(r + max_row, c = 1).value = ''
                
    workbook.save(filename=save_file)
    print('finished')